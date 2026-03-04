# engine.py

import io
import re
import zipfile
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

SMALL_TO_THOUSAND_THRESHOLD = 1_000_000
AUTO_MULTIPLIER_FOR_SMALL = 1000


@dataclass
class ChangeRow:
    file: str
    excel_row: int
    sku_full: str
    old_value: int
    new_value: int
    note: str


def _norm_str(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def normalize_header(x) -> str:
    return _norm_str(x).replace("\n", " ").replace("\r", " ").strip().upper()


def safe_cell_value(cell):
    if isinstance(cell, MergedCell):
        return None
    return cell.value


def parse_int_maybe(v) -> Optional[int]:
    if v is None:
        return None
    s = _norm_str(v)
    if s == "":
        return None
    s = s.replace(".", "").replace(",", "")
    try:
        return int(float(s))
    except Exception:
        return None


def apply_multiplier_if_needed(val: Optional[int]) -> Optional[int]:
    if val is None:
        return None
    if val < SMALL_TO_THOUSAND_THRESHOLD:
        return int(val) * AUTO_MULTIPLIER_FOR_SMALL
    return int(val)


def excel_col(letter: str) -> int:
    letter = letter.upper().strip()
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def find_col_by_headers(ws: Worksheet, header_row: int, headers: List[str]) -> Optional[int]:
    headers_norm = {normalize_header(h) for h in headers}
    for c in range(1, ws.max_column + 1):
        v = safe_cell_value(ws.cell(row=header_row, column=c))
        if normalize_header(v) in headers_norm:
            return c
    return None


def get_sheet_by_name_case_insensitive(wb, name: str):
    want = name.strip().upper()
    for s in wb.sheetnames:
        if s.strip().upper() == want:
            return wb[s]
    raise ValueError(f"Sheet '{name}' tidak ditemukan di file.")


def split_sku_addons(sku_full: str) -> Tuple[str, List[str]]:
    parts = [p.strip() for p in _norm_str(sku_full).split("+") if p.strip()]
    if not parts:
        return "", []
    return parts[0], parts[1:]


def keep_only_rows(ws: Worksheet, data_start_row: int, keep_rows: List[int]) -> None:
    keep_set = set(keep_rows)
    for r in range(ws.max_row, data_start_row - 1, -1):
        if r not in keep_set:
            ws.delete_rows(r, 1)


def make_zip(files: List[Tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in files:
            zf.writestr(name, data)
    return buf.getvalue()


# =========================
# Read pricelist (sheet change only)
# =========================
def build_pricelist_map(
    pricelist_bytes: bytes,
    sheet_name: str,
    header_row: int,
    sku_header_candidates: List[str],
    price_col_letter: str,
) -> Dict[str, int]:
    wb = load_workbook(io.BytesIO(pricelist_bytes), data_only=True)
    ws = get_sheet_by_name_case_insensitive(wb, sheet_name)

    sku_candidates_norm = {normalize_header(x) for x in sku_header_candidates}
    sku_col = None
    for c in range(1, ws.max_column + 1):
        v = safe_cell_value(ws.cell(row=header_row, column=c))
        if normalize_header(v) in sku_candidates_norm:
            sku_col = c
            break
    if sku_col is None:
        raise ValueError("Pricelist: kolom SKU tidak ditemukan (cek header row).")

    price_col = excel_col(price_col_letter)

    out: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        sku = _norm_str(safe_cell_value(ws.cell(row=r, column=sku_col)))
        if not sku:
            continue
        pv = parse_int_maybe(safe_cell_value(ws.cell(row=r, column=price_col)))
        pv = apply_multiplier_if_needed(pv)  # ✅ auto x1000
        if pv is None:
            continue
        out[sku] = int(pv)
    return out


def build_addon_map(addon_bytes: bytes, code_candidates: List[str], price_candidates: List[str]) -> Dict[str, int]:
    wb = load_workbook(io.BytesIO(addon_bytes), data_only=True)
    ws = wb.active

    code_col = None
    price_col = None
    header_row = None

    for r in range(1, min(30, ws.max_row) + 1):
        row_map = {}
        for c in range(1, ws.max_column + 1):
            row_map[normalize_header(safe_cell_value(ws.cell(row=r, column=c)))] = c

        for cand in code_candidates:
            if normalize_header(cand) in row_map:
                code_col = row_map[normalize_header(cand)]
                break
        for cand in price_candidates:
            if normalize_header(cand) in row_map:
                price_col = row_map[normalize_header(cand)]
                break

        if code_col and price_col:
            header_row = r
            break

    if not (code_col and price_col and header_row):
        raise ValueError("Addon mapping: header tidak ditemukan (butuh addon_code & harga).")

    out: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = _norm_str(safe_cell_value(ws.cell(row=r, column=code_col)))
        if not code:
            continue
        pv = parse_int_maybe(safe_cell_value(ws.cell(row=r, column=price_col)))
        pv = apply_multiplier_if_needed(pv)  # ✅ auto x1000
        if pv is None:
            continue
        out[code] = int(pv)
    return out


# =========================
# Price in-place (Harga Normal + Shopee coret)
# =========================
def process_price_inplace(
    template_bytes: bytes,
    template_name: str,
    spec: dict,
    pricelist_map: Dict[str, int],
    addon_map: Dict[str, int],
    discount_rp: int = 0,
    only_changed: bool = True,
) -> Tuple[Optional[bytes], List[ChangeRow]]:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    header_row = spec["template"]["header_row"]
    data_start_row = spec["template"]["data_start_row"]

    sku_col = find_col_by_headers(ws, header_row, spec["template"]["sku_headers"])
    price_col = find_col_by_headers(ws, header_row, spec["template"]["price_headers"])

    if sku_col is None or price_col is None:
        raise ValueError(f"[{template_name}] kolom SKU/Harga tidak ditemukan (cek header).")

    disc = int(discount_rp or 0)
    if disc < 0:
        disc = 0

    changed_rows: List[int] = []
    changes: List[ChangeRow] = []

    for r in range(data_start_row, ws.max_row + 1):
        sku_full = _norm_str(safe_cell_value(ws.cell(row=r, column=sku_col)))
        if not sku_full:
            continue

        old_val = parse_int_maybe(safe_cell_value(ws.cell(row=r, column=price_col))) or 0

        base, addons = split_sku_addons(sku_full)
        if not base or base not in pricelist_map:
            continue  # skip

        total = int(pricelist_map[base])

        ok = True
        for a in addons:
            if a not in addon_map:
                ok = False
                break
            total += int(addon_map[a])
        if not ok:
            continue  # skip

        total = total - disc
        if total < 0:
            total = 0

        if only_changed and int(total) == int(old_val):
            continue

        ws.cell(row=r, column=price_col).value = int(total)

        changed_rows.append(r)
        changes.append(ChangeRow(template_name, r, sku_full, int(old_val), int(total), "changed"))

    if not changed_rows:
        return None, changes

    keep_only_rows(ws, data_start_row, changed_rows)
    return workbook_to_bytes(wb), changes


# =========================
# Stock: read from sheet range LAPTOP..SER OTH CON
# + allow choose column (Nasional/Area/Toko)
# =========================
def iter_sheets_range(wb, from_name: str, to_name: str) -> List[str]:
    names = wb.sheetnames
    up = [n.strip().upper() for n in names]
    f = from_name.strip().upper()
    t = to_name.strip().upper()

    if f not in up or t not in up:
        raise ValueError(f"Sheet range '{from_name}'..'{to_name}' tidak ditemukan di workbook.")

    i1 = up.index(f)
    i2 = up.index(t)
    if i1 > i2:
        i1, i2 = i2, i1
    return names[i1 : i2 + 1]


def build_stock_dataframe_from_range(stock_file_bytes: bytes, sheets_from: str, sheets_to: str) -> pd.DataFrame:
    wb = load_workbook(io.BytesIO(stock_file_bytes), data_only=True)
    sheet_names = iter_sheets_range(wb, sheets_from, sheets_to)

    combined_rows = []
    for sname in sheet_names:
        ws = wb[sname]

        header_row = None
        header_map = None

        # scan header 1..15
        for r in range(1, min(15, ws.max_row) + 1):
            m = {}
            for c in range(1, ws.max_column + 1):
                key = normalize_header(ws.cell(r, c).value)
                if key:
                    m[key] = c

            if "KODEBARANG" in m or "KODE BARANG" in m or "SKU" in m:
                header_row = r
                header_map = m
                break

        if header_row is None or header_map is None:
            continue

        cols = list(header_map.items())
        for rr in range(header_row + 1, ws.max_row + 1):
            row_dict = {}
            for name_norm, col_idx in cols:
                row_dict[name_norm] = ws.cell(rr, col_idx).value
            combined_rows.append(row_dict)

    if not combined_rows:
        raise ValueError("File stok: tidak ada data terbaca dari sheet range (LAPTOP..SER OTH CON).")

    return pd.DataFrame(combined_rows)


def build_stock_map_from_df(df: pd.DataFrame, qty_col: str) -> Dict[str, int]:
    sku_col = None
    for k in ["KODEBARANG", "KODE BARANG", "SKU"]:
        if k in df.columns:
            sku_col = k
            break
    if sku_col is None:
        raise ValueError("File stok: kolom SKU (KODEBARANG/KODE BARANG/SKU) tidak ditemukan.")

    if qty_col not in df.columns:
        raise ValueError(f"Kolom stok '{qty_col}' tidak ditemukan.")

    out: Dict[str, int] = {}
    for _, row in df.iterrows():
        sku = _norm_str(row.get(sku_col))
        if not sku:
            continue
        sku_key = re.sub(r"\s+", "", sku.strip().upper())
        v = row.get(qty_col)
        try:
            qty = int(float(v))
        except Exception:
            continue
        out[sku_key] = qty
    return out


def process_stock_inplace(
    template_bytes: bytes,
    template_name: str,
    spec: dict,
    stock_value_map: Dict[str, int],
) -> Tuple[Optional[bytes], List[ChangeRow]]:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    header_row = spec["template"]["header_row"]
    data_start_row = spec["template"]["data_start_row"]

    sku_col = find_col_by_headers(ws, header_row, spec["template"]["sku_headers"])
    qty_col = find_col_by_headers(ws, header_row, spec["template"]["stock_headers"])

    if sku_col is None or qty_col is None:
        raise ValueError(f"[{template_name}] kolom SKU/Stok tidak ditemukan (cek header).")

    def norm_sku(v) -> str:
        s = _norm_str(v).upper()
        if re.fullmatch(r"\d+\.0", s):
            s = s[:-2]
        s = re.sub(r"\s+", "", s)
        return s

    changed_rows: List[int] = []
    changes: List[ChangeRow] = []

    for r in range(data_start_row, ws.max_row + 1):
        sku_full = _norm_str(safe_cell_value(ws.cell(row=r, column=sku_col)))
        if not sku_full:
            continue

        base, _addons = split_sku_addons(sku_full)
        key = norm_sku(base)
        if not key:
            continue

        if key not in stock_value_map:
            continue

        old_val = parse_int_maybe(safe_cell_value(ws.cell(row=r, column=qty_col))) or 0
        new_val = int(stock_value_map[key])

        if int(old_val) == int(new_val):
            continue

        ws.cell(row=r, column=qty_col).value = int(new_val)
        changed_rows.append(r)
        changes.append(ChangeRow(template_name, r, sku_full, int(old_val), int(new_val), "changed"))

    if not changed_rows:
        return None, changes

    keep_only_rows(ws, data_start_row, changed_rows)
    return workbook_to_bytes(wb), changes


# =========================
# Discount Template output (TikTok & PM) split 1000
# =========================
def chunk_list(items: List[dict], size: int) -> List[List[dict]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def build_discount_template_workbook(rows: List[dict], headers: List[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i).value = h

    r = 2
    for it in rows:
        ws.cell(row=r, column=1).value = it.get("product_id", "")
        ws.cell(row=r, column=2).value = it.get("sku_id", "")
        ws.cell(row=r, column=3).value = it.get("offer_price", "")
        ws.cell(row=r, column=4).value = it.get("promo_stock", "")
        r += 1

    return workbook_to_bytes(wb)


def process_discount_template(
    template_bytes: bytes,
    template_name: str,
    spec: dict,
    pricelist_map: Dict[str, int],
    addon_map: Dict[str, int],
    discount_rp: int,
    only_changed: bool = True,
) -> Tuple[List[Tuple[str, bytes]], pd.DataFrame]:
    wb = load_workbook(io.BytesIO(template_bytes), data_only=True)
    ws = wb.active

    data_start_row = spec["input"]["data_start_row"]

    col_product_id = excel_col(spec["input"]["col_product_id"])
    col_sku_id = excel_col(spec["input"]["col_sku_id"])
    col_price = excel_col(spec["input"]["col_price"])
    col_stock = excel_col(spec["input"]["col_stock"])
    col_seller_sku = excel_col(spec["input"]["col_seller_sku"])

    def col_is_all_empty(col_idx: int) -> bool:
        for rr in range(data_start_row, min(ws.max_row, data_start_row + 50) + 1):
            v = ws.cell(row=rr, column=col_idx).value
            if v is not None and _norm_str(v) != "":
                return False
        return True

    if col_is_all_empty(col_seller_sku):
        col_seller_sku = excel_col("E")

    disc = int(discount_rp or 0)
    if disc < 0:
        disc = 0

    out_rows: List[dict] = []
    preview_rows: List[dict] = []

    for r in range(data_start_row, ws.max_row + 1):
        product_id = _norm_str(ws.cell(row=r, column=col_product_id).value)
        sku_id = _norm_str(ws.cell(row=r, column=col_sku_id).value)

        old_price = parse_int_maybe(ws.cell(row=r, column=col_price).value) or 0
        promo_stock = parse_int_maybe(ws.cell(row=r, column=col_stock).value)
        promo_stock = int(promo_stock) if promo_stock is not None else ""

        seller_sku_full = _norm_str(ws.cell(row=r, column=col_seller_sku).value)

        if not product_id and not sku_id and not seller_sku_full:
            continue

        base, addons = split_sku_addons(seller_sku_full)
        if not base or base not in pricelist_map:
            continue

        total = int(pricelist_map[base])
        ok = True
        for a in addons:
            if a not in addon_map:
                ok = False
                break
            total += int(addon_map[a])
        if not ok:
            continue

        new_offer = total - disc
        if new_offer < 0:
            new_offer = 0

        if only_changed and int(new_offer) == int(old_price):
            continue

        out_rows.append({
            "product_id": product_id,
            "sku_id": sku_id,
            "offer_price": int(new_offer),
            "promo_stock": promo_stock,
        })

        preview_rows.append({
            "template": template_name,
            "row": r,
            "seller_sku": seller_sku_full,
            "old_price": int(old_price),
            "new_offer_price": int(new_offer),
        })

    headers = spec["output"]["headers"]
    max_rows = int(spec["output"]["max_rows_per_file"])

    chunks = chunk_list(out_rows, max_rows)
    out_files: List[Tuple[str, bytes]] = []

    if not chunks or (len(chunks) == 1 and len(chunks[0]) == 0):
        return [], pd.DataFrame(preview_rows)

    if len(chunks) == 1:
        out_xlsx = build_discount_template_workbook(chunks[0], headers)
        out_files.append((f"{template_name.replace('.xlsx','')}_Product Discount.xlsx", out_xlsx))
    else:
        for i, ch in enumerate(chunks, start=1):
            out_xlsx = build_discount_template_workbook(ch, headers)
            out_files.append((f"{template_name.replace('.xlsx','')}_Product Discount {i}.xlsx", out_xlsx))

    return out_files, pd.DataFrame(preview_rows)
