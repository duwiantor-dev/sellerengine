# app.py  (FULL - single file, with DEBUG MODE)
import io
import re
import zipfile
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

# =========================
# STREAMLIT CONFIG
# =========================
st.set_page_config(page_title="sellerengine", page_icon="⚙️", layout="wide")

# Compact UI CSS (kecilkan uploader & input)
st.markdown(
    """
<style>
[data-testid="stFileUploaderDropzone"] { min-height: 64px; padding: 8px 10px; }
[data-testid="stFileUploaderDropzone"] * { font-size: 12px; line-height: 1.2; }
div[data-testid="stNumberInput"] input { padding-top: 6px; padding-bottom: 6px; }
</style>
""",
    unsafe_allow_html=True,
)

# =========================
# CONFIG: pricelist tanpa 000 => auto x1000
# =========================
SMALL_TO_THOUSAND_THRESHOLD = 1_000_000
AUTO_MULTIPLIER_FOR_SMALL = 1000

# =========================
# SPECS (RULE PER MARKETPLACE)
# =========================
SPECS = {
    # -------------------------
    # HARGA NORMAL
    # -------------------------
    ("harga_normal", "tiktok"): {
        "template": {
            "header_row": 3,
            "data_start_row": 6,
            "sku_headers": ["SKU Penjual", "Seller SKU"],
            "price_headers": ["Harga Ritel (Mata Uang Lokal)"],
        },
        "pricelist": {
            "sheet_name": "CHANGE",
            "header_row": 2,
            "sku_header_candidates": ["KODEBARANG", "KODE BARANG"],
            "price_col_letter": "M3",
        },
        "addon": {
            "code_candidates": ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON"],
            "price_candidates": ["harga", "HARGA", "Price", "PRICE", "Harga"],
        },
    },
    ("harga_normal", "shopee"): {
        "template": {
            "header_row": 3,
            "data_start_row": 7,
            "sku_headers": ["SKU", "SKU Ref. No.(Optional)", "SKU\u00a0Ref.\u00a0No.(Optional)"],
            "price_headers": ["Harga", "Price", "Harga Normal", "Harga Ritel (Mata Uang Lokal)"],
        },
        "pricelist": {
            "sheet_name": "CHANGE",
            "header_row": 2,
            "sku_header_candidates": ["KODEBARANG", "KODE BARANG"],
            "price_col_letter": "M4",
        },
        "addon": {
            "code_candidates": ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON"],
            "price_candidates": ["harga", "HARGA", "Price", "PRICE", "Harga"],
        },
    },
    ("harga_normal", "powermerchant"): {
        "template": {
            "header_row": 3,
            "data_start_row": 6,
            "sku_headers": ["SKU Penjual", "Seller SKU"],
            "price_headers": ["Harga Ritel (Mata Uang Lokal)"],
        },
        "pricelist": {
            "sheet_name": "CHANGE",
            "header_row": 2,
            "sku_header_candidates": ["KODEBARANG", "KODE BARANG"],
            "price_col_letter": "M4",
        },
        "addon": {
            "code_candidates": ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON"],
            "price_candidates": ["harga", "HARGA", "Price", "PRICE", "Harga"],
        },
    },
    ("harga_normal", "bigseller"): {
        "template": {
            "header_row": 1,
            "data_start_row": 2,
            "sku_headers": ["SKU"],     # ✅ BigSeller SKU
            "price_headers": ["Harga"], # ✅ BigSeller Harga
        },
        "pricelist": {
            "sheet_name": "CHANGE",
            "header_row": 2,
            "sku_header_candidates": ["KODEBARANG", "KODE BARANG"],
            "price_col_letter": "M4",
        },
        "addon": {
            "code_candidates": ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON"],
            "price_candidates": ["harga", "HARGA", "Price", "PRICE", "Harga"],
        },
    },

    # -------------------------
    # HARGA CORET (SHOPEE IN-PLACE)
    # -------------------------
    ("harga_coret", "shopee"): {
        "template": {
            "header_row": 1,
            "data_start_row": 2,
            "sku_headers": ["SKU Ref. No.(Optional)", "SKU\u00a0Ref.\u00a0No.(Optional)", "SKU"],
            "price_headers": ["Harga diskon", "Discount Price", "Harga Diskon"],
        },
        "pricelist": {
            "sheet_name": "CHANGE",
            "header_row": 2,
            "sku_header_candidates": ["KODEBARANG", "KODE BARANG"],
            "price_col_letter": "M4",
        },
        "addon": {
            "code_candidates": ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON"],
            "price_candidates": ["harga", "HARGA", "Price", "PRICE", "Harga"],
        },
    },

    # -------------------------
    # DISCOUNT TEMPLATE (TIKTOK & POWERMERCHANT) + split 1000
    # -------------------------
    ("discount_template", "tiktok"): {
        "input": {
            "data_start_row": 6,
            "col_product_id": "A",
            "col_sku_id": "D",
            "col_price": "F",
            "col_stock": "G",
            "col_seller_sku": "H",  # fallback E jika kosong
        },
        "output": {
            "max_rows_per_file": 1000,
            "headers": [
                "Product_id (wajib)",
                "SKU_id (wajib)",
                "Harga Penawaran (wajib)",
                "Total Stok Promosi (opsional)\n1. Total Stok Promosi≤ Stok \n2. Jika tidak diisi artinya tidak terbatas",
                "Batas Pembelian (opsional)\n1. 1 ≤ Batas pembelian≤ 99\n2. Jika tidak diisi artinya tidak terbatas",
            ],
        },
        "pricelist": {
            "sheet_name": "CHANGE",
            "header_row": 2,
            "sku_header_candidates": ["KODEBARANG", "KODE BARANG"],
            "price_col_letter": "M3",
        },
        "addon": {
            "code_candidates": ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON"],
            "price_candidates": ["harga", "HARGA", "Price", "PRICE", "Harga"],
        },
    },
    ("discount_template", "powermerchant"): {
        "input": {
            "data_start_row": 6,
            "col_product_id": "A",
            "col_sku_id": "D",
            "col_price": "F",
            "col_stock": "G",
            "col_seller_sku": "H",
        },
        "output": {
            "max_rows_per_file": 1000,
            "headers": [
                "Product_id (wajib)",
                "SKU_id (wajib)",
                "Harga Penawaran (wajib)",
                "Total Stok Promosi (opsional)\n1. Total Stok Promosi≤ Stok \n2. Jika tidak diisi artinya tidak terbatas",
                "Batas Pembelian (opsional)\n1. 1 ≤ Batas pembelian≤ 99\n2. Jika tidak diisi artinya tidak terbatas",
            ],
        },
        "pricelist": {
            "sheet_name": "CHANGE",
            "header_row": 2,
            "sku_header_candidates": ["KODEBARANG", "KODE BARANG"],
            "price_col_letter": "M4",  # ✅ PM = M4
        },
        "addon": {
            "code_candidates": ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON"],
            "price_candidates": ["harga", "HARGA", "Price", "PRICE", "Harga"],
        },
    },

    # -------------------------
    # UPDATE STOK
    # -------------------------
    ("update_stok", "tiktok"): {
        "template": {
            "header_row": 3,
            "data_start_row": 6,
            "sku_headers": ["SKU Penjual", "Seller SKU"],
            "stock_headers": ["Kuantitas", "Quantity", "Qty"],
        },
        "stock_source": {"sheets_from": "LAPTOP", "sheets_to": "SER OTH CON"},
    },
    ("update_stok", "shopee"): {
        "template": {
            "header_row": 3,
            "data_start_row": 7,
            "sku_headers": ["SKU"],
            "stock_headers": ["Stok", "Stock"],
        },
        "stock_source": {"sheets_from": "LAPTOP", "sheets_to": "SER OTH CON"},
    },
}

# =========================
# DATA STRUCT
# =========================
@dataclass
class ChangeRow:
    file: str
    excel_row: int
    sku_full: str
    old_value: int
    new_value: int
    note: str


# =========================
# UTIL
# =========================
def _norm_str(x) -> str:
    if x is None:
        return ""
    return str(x).strip()

def normalize_header(x) -> str:
    return _norm_str(x).replace("\n", " ").replace("\r", " ").strip().upper()

def safe_cell_value(cell):
    if isinstance(cell, MergedCell):
        return None
    return cell.value

def parse_int_maybe(v) -> Optional[int]:
    if v is None:
        return None
    s = _norm_str(v)
    if s == "":
        return None
    s = s.replace(".", "").replace(",", "")
    try:
        return int(float(s))
    except Exception:
        return None

def apply_multiplier_if_needed(val: Optional[int]) -> Optional[int]:
    if val is None:
        return None
    if val < SMALL_TO_THOUSAND_THRESHOLD:
        return int(val) * AUTO_MULTIPLIER_FOR_SMALL
    return int(val)

def excel_col(letter: str) -> int:
    letter = letter.upper().strip()
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n

def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def find_col_by_headers(ws: Worksheet, header_row: int, headers: List[str]) -> Optional[int]:
    headers_norm = {normalize_header(h) for h in headers}
    for c in range(1, ws.max_column + 1):
        v = safe_cell_value(ws.cell(row=header_row, column=c))
        if normalize_header(v) in headers_norm:
            return c
    return None

def get_sheet_by_name_case_insensitive(wb, name: str):
    want = name.strip().upper()
    for s in wb.sheetnames:
        if s.strip().upper() == want:
            return wb[s]
    raise ValueError(f"Sheet '{name}' tidak ditemukan di file.")

def split_sku_addons(sku_full: str) -> Tuple[str, List[str]]:
    parts = [p.strip() for p in _norm_str(sku_full).split("+") if p.strip()]
    if not parts:
        return "", []
    return parts[0], parts[1:]

def norm_sku_key(v: str) -> str:
    """Normalisasi SKU supaya match lebih sering: trim, uppercase, hapus spasi, ubah '123.0' => '123'."""
    s = _norm_str(v).upper()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    s = re.sub(r"\s+", "", s)
    return s

def keep_only_rows(ws: Worksheet, data_start_row: int, keep_rows: List[int]) -> None:
    keep_set = set(keep_rows)
    for r in range(ws.max_row, data_start_row - 1, -1):
        if r not in keep_set:
            ws.delete_rows(r, 1)

def make_zip(files: List[Tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in files:
            zf.writestr(name, data)
    return buf.getvalue()

# =========================
# CACHE: build pricelist/addon map (biar cepat)
# =========================
@st.cache_data(show_spinner=False)
def cached_build_pricelist_map(
    pricelist_bytes: bytes,
    sheet_name: str,
    header_row: int,
    sku_header_candidates: Tuple[str, ...],
    price_col_letter: str,
) -> Dict[str, int]:
    wb = load_workbook(io.BytesIO(pricelist_bytes), data_only=True)
    ws = get_sheet_by_name_case_insensitive(wb, sheet_name)

    sku_candidates_norm = {normalize_header(x) for x in sku_header_candidates}

    sku_col = None
    for c in range(1, ws.max_column + 1):
        v = safe_cell_value(ws.cell(row=header_row, column=c))
        if normalize_header(v) in sku_candidates_norm:
            sku_col = c
            break
    if sku_col is None:
        raise ValueError("Pricelist: kolom SKU tidak ditemukan (cek header row).")

    price_col = excel_col(price_col_letter)

    out: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):

    sku_no = safe_cell_value(ws.cell(row=r, column=1))
    if sku_no is None:
        continue

    sku_raw = _norm_str(safe_cell_value(ws.cell(row=r, column=sku_col)))
    if not sku_raw:
        continue

    sku = norm_sku_key(sku_raw)

    pv = parse_int_maybe(safe_cell_value(ws.cell(row=r, column=price_col)))
    pv = apply_multiplier_if_needed(pv)

    if pv is None:
        continue

    out[sku] = int(pv)
    return out

@st.cache_data(show_spinner=False)
def cached_build_addon_map(
    addon_bytes: bytes,
    code_candidates: Tuple[str, ...],
    price_candidates: Tuple[str, ...],
) -> Dict[str, int]:
    wb = load_workbook(io.BytesIO(addon_bytes), data_only=True)
    ws = wb.active

    code_col = None
    price_col = None
    header_row = None

    for r in range(1, min(30, ws.max_row) + 1):
        row_map = {}
        for c in range(1, ws.max_column + 1):
            row_map[normalize_header(safe_cell_value(ws.cell(row=r, column=c)))] = c

        for cand in code_candidates:
            if normalize_header(cand) in row_map:
                code_col = row_map[normalize_header(cand)]
                break
        for cand in price_candidates:
            if normalize_header(cand) in row_map:
                price_col = row_map[normalize_header(cand)]
                break

        if code_col and price_col:
            header_row = r
            break

    if not (code_col and price_col and header_row):
        raise ValueError("Addon mapping: header tidak ditemukan (butuh addon_code & harga).")

    out: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):

    sku_no = safe_cell_value(ws.cell(row=r, column=1))
    if sku_no is None:
        continue

    sku_raw = _norm_str(safe_cell_value(ws.cell(row=r, column=sku_col)))
    if not sku_raw:
        continue

    sku = norm_sku_key(sku_raw)

    pv = parse_int_maybe(safe_cell_value(ws.cell(row=r, column=price_col)))
    pv = apply_multiplier_if_needed(pv)

    if pv is None:
        continue

    out[sku] = int(pv)
    return out

# =========================
# ENGINE: PRICE IN-PLACE (with debug info)
# =========================
def process_price_inplace(
    template_bytes: bytes,
    template_name: str,
    spec: dict,
    pricelist_map: Dict[str, int],
    addon_map: Dict[str, int],
    discount_rp: int = 0,   # diskon manual: rupiah full, TIDAK x1000
    only_changed: bool = True,
    debug: bool = False,
    debug_limit: int = 50,
) -> Tuple[Optional[bytes], List[ChangeRow], Dict]:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    header_row = spec["template"]["header_row"]
    data_start_row = spec["template"]["data_start_row"]

    sku_col = find_col_by_headers(ws, header_row, spec["template"]["sku_headers"])
    price_col = find_col_by_headers(ws, header_row, spec["template"]["price_headers"])

    if sku_col is None or price_col is None:
        raise ValueError(f"[{template_name}] kolom SKU/Harga tidak ditemukan (cek header).")

    disc = int(discount_rp or 0)
    if disc < 0:
        disc = 0

    changed_rows: List[int] = []
    changes: List[ChangeRow] = []

    dbg = {
        "template": template_name,
        "rows_total_scanned": 0,
        "rows_with_sku": 0,
        "rows_base_found": 0,
        "rows_addon_ok": 0,
        "rows_changed": 0,
        "rows_same_price": 0,
        "rows_skipped_base_not_found": 0,
        "rows_skipped_addon_not_found": 0,
        "sample_missing_base": [],
        "sample_missing_addon": [],
        "sample_first_sku": [],
    }

    for r in range(data_start_row, ws.max_row + 1):
        dbg["rows_total_scanned"] += 1

        sku_full_raw = _norm_str(safe_cell_value(ws.cell(row=r, column=sku_col)))
        if not sku_full_raw:
            continue

        dbg["rows_with_sku"] += 1
        if debug and len(dbg["sample_first_sku"]) < 10:
            dbg["sample_first_sku"].append(sku_full_raw)

        old_val = parse_int_maybe(safe_cell_value(ws.cell(row=r, column=price_col))) or 0

        base_raw, addons_raw = split_sku_addons(sku_full_raw)
        base = norm_sku_key(base_raw)
        addons = [norm_sku_key(a) for a in addons_raw]

        if not base or base not in pricelist_map:
            dbg["rows_skipped_base_not_found"] += 1
            if debug and len(dbg["sample_missing_base"]) < debug_limit and base_raw:
                dbg["sample_missing_base"].append({"row": r, "base": base_raw})
            continue

        dbg["rows_base_found"] += 1
        total = int(pricelist_map[base])

        ok = True
        for a in addons:
            if a and a not in addon_map:
                ok = False
                dbg["rows_skipped_addon_not_found"] += 1
                if debug and len(dbg["sample_missing_addon"]) < debug_limit:
                    dbg["sample_missing_addon"].append({"row": r, "sku": sku_full_raw, "missing_addon": a})
                break
            if a:
                total += int(addon_map[a])

        if not ok:
            continue

        dbg["rows_addon_ok"] += 1

        total = total - disc
        if total < 0:
            total = 0

        if only_changed and int(total) == int(old_val):
            dbg["rows_same_price"] += 1
            continue

        ws.cell(row=r, column=price_col).value = int(total)
        changed_rows.append(r)
        changes.append(ChangeRow(template_name, r, sku_full_raw, int(old_val), int(total), "changed"))
        dbg["rows_changed"] += 1

    if not changed_rows:
        return None, changes, dbg

    keep_only_rows(ws, data_start_row, changed_rows)
    return workbook_to_bytes(wb), changes, dbg

# =========================
# ENGINE: DISCOUNT TEMPLATE (TikTok/PM) split 1000
# =========================
def chunk_list(items: List[dict], size: int) -> List[List[dict]]:
    return [items[i:i + size] for i in range(0, len(items), size)]

def build_discount_template_workbook(rows: List[dict], headers: List[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i).value = h

    r = 2
    for it in rows:
        ws.cell(row=r, column=1).value = it.get("product_id", "")
        ws.cell(row=r, column=2).value = it.get("sku_id", "")
        ws.cell(row=r, column=3).value = it.get("offer_price", "")
        ws.cell(row=r, column=4).value = it.get("promo_stock", "")
        r += 1

    return workbook_to_bytes(wb)

def process_discount_template(
    template_bytes: bytes,
    template_name: str,
    spec: dict,
    pricelist_map: Dict[str, int],
    addon_map: Dict[str, int],
    discount_rp: int,
    only_changed: bool = True,
    debug: bool = False,
    debug_limit: int = 50,
) -> Tuple[List[Tuple[str, bytes]], pd.DataFrame, Dict]:
    wb = load_workbook(io.BytesIO(template_bytes), data_only=True)
    ws = wb.active

    data_start_row = spec["input"]["data_start_row"]
    col_product_id = excel_col(spec["input"]["col_product_id"])
    col_sku_id = excel_col(spec["input"]["col_sku_id"])
    col_price = excel_col(spec["input"]["col_price"])
    col_stock = excel_col(spec["input"]["col_stock"])
    col_seller_sku = excel_col(spec["input"]["col_seller_sku"])

    def col_is_all_empty(col_idx: int) -> bool:
        for rr in range(data_start_row, min(ws.max_row, data_start_row + 50) + 1):
            v = ws.cell(row=rr, column=col_idx).value
            if v is not None and _norm_str(v) != "":
                return False
        return True

    if col_is_all_empty(col_seller_sku):
        col_seller_sku = excel_col("E")

    disc = int(discount_rp or 0)
    if disc < 0:
        disc = 0

    out_rows: List[dict] = []
    preview_rows: List[dict] = []

    dbg = {
        "template": template_name,
        "rows_total_scanned": 0,
        "rows_with_seller_sku": 0,
        "rows_base_found": 0,
        "rows_addon_ok": 0,
        "rows_changed": 0,
        "rows_same_price": 0,
        "rows_skipped_base_not_found": 0,
        "rows_skipped_addon_not_found": 0,
        "sample_missing_base": [],
        "sample_missing_addon": [],
        "sample_first_sku": [],
    }

    for r in range(data_start_row, ws.max_row + 1):
        dbg["rows_total_scanned"] += 1

        product_id = _norm_str(ws.cell(row=r, column=col_product_id).value)
        sku_id = _norm_str(ws.cell(row=r, column=col_sku_id).value)
        old_price = parse_int_maybe(ws.cell(row=r, column=col_price).value) or 0
        promo_stock = parse_int_maybe(ws.cell(row=r, column=col_stock).value)
        promo_stock = int(promo_stock) if promo_stock is not None else ""

        seller_sku_full_raw = _norm_str(ws.cell(row=r, column=col_seller_sku).value)
        if not product_id and not sku_id and not seller_sku_full_raw:
            continue
        if not seller_sku_full_raw:
            continue

        dbg["rows_with_seller_sku"] += 1
        if debug and len(dbg["sample_first_sku"]) < 10:
            dbg["sample_first_sku"].append(seller_sku_full_raw)

        base_raw, addons_raw = split_sku_addons(seller_sku_full_raw)
        base = norm_sku_key(base_raw)
        addons = [norm_sku_key(a) for a in addons_raw]

        if not base or base not in pricelist_map:
            dbg["rows_skipped_base_not_found"] += 1
            if debug and len(dbg["sample_missing_base"]) < debug_limit and base_raw:
                dbg["sample_missing_base"].append({"row": r, "base": base_raw})
            continue

        dbg["rows_base_found"] += 1
        total = int(pricelist_map[base])

        ok = True
        for a in addons:
            if a and a not in addon_map:
                ok = False
                dbg["rows_skipped_addon_not_found"] += 1
                if debug and len(dbg["sample_missing_addon"]) < debug_limit:
                    dbg["sample_missing_addon"].append({"row": r, "sku": seller_sku_full_raw, "missing_addon": a})
                break
            if a:
                total += int(addon_map[a])

        if not ok:
            continue

        dbg["rows_addon_ok"] += 1

        new_offer = total - disc
        if new_offer < 0:
            new_offer = 0

        if only_changed and int(new_offer) == int(old_price):
            dbg["rows_same_price"] += 1
            continue

        out_rows.append({
            "product_id": product_id,
            "sku_id": sku_id,
            "offer_price": int(new_offer),
            "promo_stock": promo_stock,
        })

        preview_rows.append({
            "template": template_name,
            "row": r,
            "seller_sku": seller_sku_full_raw,
            "old_price": int(old_price),
            "new_offer_price": int(new_offer),
        })
        dbg["rows_changed"] += 1

    headers = spec["output"]["headers"]
    max_rows = int(spec["output"]["max_rows_per_file"])
    chunks = chunk_list(out_rows, max_rows)
    out_files: List[Tuple[str, bytes]] = []

    if not chunks or (len(chunks) == 1 and len(chunks[0]) == 0):
        return [], pd.DataFrame(preview_rows), dbg

    if len(chunks) == 1:
        out_xlsx = build_discount_template_workbook(chunks[0], headers)
        out_files.append((f"{template_name.replace('.xlsx','')}_Product Discount.xlsx", out_xlsx))
    else:
        for i, ch in enumerate(chunks, start=1):
            out_xlsx = build_discount_template_workbook(ch, headers)
            out_files.append((f"{template_name.replace('.xlsx','')}_Product Discount {i}.xlsx", out_xlsx))

    return out_files, pd.DataFrame(preview_rows), dbg

# =========================
# ENGINE: STOCK
# =========================
def iter_sheets_range(wb, from_name: str, to_name: str) -> List[str]:
    names = wb.sheetnames
    up = [n.strip().upper() for n in names]
    f = from_name.strip().upper()
    t = to_name.strip().upper()

    if f not in up or t not in up:
        raise ValueError(f"Sheet range '{from_name}'..'{to_name}' tidak ditemukan di workbook.")

    i1 = up.index(f)
    i2 = up.index(t)
    if i1 > i2:
        i1, i2 = i2, i1
    return names[i1:i2 + 1]

@st.cache_data(show_spinner=False)
def cached_build_stock_dataframe_from_range(stock_file_bytes: bytes, sheets_from: str, sheets_to: str) -> pd.DataFrame:
    wb = load_workbook(io.BytesIO(stock_file_bytes), data_only=True)
    sheet_names = iter_sheets_range(wb, sheets_from, sheets_to)

    combined_rows = []
    for sname in sheet_names:
        ws = wb[sname]

        header_row = None
        header_map = None

        for r in range(1, min(15, ws.max_row) + 1):
            m = {}
            for c in range(1, ws.max_column + 1):
                key = normalize_header(ws.cell(r, c).value)
                if key:
                    m[key] = c

            if "KODEBARANG" in m or "KODE BARANG" in m or "SKU" in m:
                header_row = r
                header_map = m
                break

        if header_row is None or header_map is None:
            continue

        cols = list(header_map.items())
        for rr in range(header_row + 1, ws.max_row + 1):
            row_dict = {}
            for name_norm, col_idx in cols:
                row_dict[name_norm] = ws.cell(rr, col_idx).value
            combined_rows.append(row_dict)

    if not combined_rows:
        raise ValueError("File stok: tidak ada data terbaca dari sheet range (LAPTOP..SER OTH CON).")

    return pd.DataFrame(combined_rows)

def build_stock_map_from_df(df: pd.DataFrame, qty_col: str) -> Dict[str, int]:
    sku_col = None
    for k in ["KODEBARANG", "KODE BARANG", "SKU"]:
        if k in df.columns:
            sku_col = k
            break
    if sku_col is None:
        raise ValueError("File stok: kolom SKU (KODEBARANG/KODE BARANG/SKU) tidak ditemukan.")

    if qty_col not in df.columns:
        raise ValueError(f"Kolom stok '{qty_col}' tidak ditemukan.")

    out: Dict[str, int] = {}
    for _, row in df.iterrows():
        sku = _norm_str(row.get(sku_col))
        if not sku:
            continue
        sku_key = norm_sku_key(sku)
        v = row.get(qty_col)
        try:
            qty = int(float(v))
        except Exception:
            continue
        out[sku_key] = qty
    return out

def process_stock_inplace(
    template_bytes: bytes,
    template_name: str,
    spec: dict,
    stock_value_map: Dict[str, int],
    debug: bool = False,
    debug_limit: int = 50,
) -> Tuple[Optional[bytes], List[ChangeRow], Dict]:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    header_row = spec["template"]["header_row"]
    data_start_row = spec["template"]["data_start_row"]

    sku_col = find_col_by_headers(ws, header_row, spec["template"]["sku_headers"])
    qty_col = find_col_by_headers(ws, header_row, spec["template"]["stock_headers"])

    if sku_col is None or qty_col is None:
        raise ValueError(f"[{template_name}] kolom SKU/Stok tidak ditemukan (cek header).")

    changed_rows: List[int] = []
    changes: List[ChangeRow] = []

    dbg = {
        "template": template_name,
        "rows_total_scanned": 0,
        "rows_with_sku": 0,
        "rows_stock_found": 0,
        "rows_changed": 0,
        "rows_same_qty": 0,
        "rows_skipped_stock_not_found": 0,
        "sample_missing_stock": [],
        "sample_first_sku": [],
    }

    for r in range(data_start_row, ws.max_row + 1):
        dbg["rows_total_scanned"] += 1
        sku_full_raw = _norm_str(safe_cell_value(ws.cell(row=r, column=sku_col)))
        if not sku_full_raw:
            continue

        dbg["rows_with_sku"] += 1
        if debug and len(dbg["sample_first_sku"]) < 10:
            dbg["sample_first_sku"].append(sku_full_raw)

        base_raw, _addons = split_sku_addons(sku_full_raw)
        key = norm_sku_key(base_raw)
        if not key:
            continue

        if key not in stock_value_map:
            dbg["rows_skipped_stock_not_found"] += 1
            if debug and len(dbg["sample_missing_stock"]) < debug_limit:
                dbg["sample_missing_stock"].append({"row": r, "base": base_raw})
            continue

        dbg["rows_stock_found"] += 1

        old_val = parse_int_maybe(safe_cell_value(ws.cell(row=r, column=qty_col))) or 0
        new_val = int(stock_value_map[key])

        if int(old_val) == int(new_val):
            dbg["rows_same_qty"] += 1
            continue

        ws.cell(row=r, column=qty_col).value = int(new_val)
        changed_rows.append(r)
        changes.append(ChangeRow(template_name, r, sku_full_raw, int(old_val), int(new_val), "changed"))
        dbg["rows_changed"] += 1

    if not changed_rows:
        return None, changes, dbg

    keep_only_rows(ws, data_start_row, changed_rows)
    return workbook_to_bytes(wb), changes, dbg


# =========================
# UI HELPERS
# =========================
def download_outputs(out_files, zip_name: str):
    if not out_files:
        st.warning("Tidak ada baris yang berubah / semua baris skip.")
        return

    if len(out_files) == 1:
        name, data = out_files[0]
        st.download_button(
            "Download XLSX",
            data=data,
            file_name=name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        z = make_zip(out_files)
        st.download_button("Download ZIP", data=z, file_name=zip_name, mime="application/zip")

def ui_row_4(prefix: str, label_mass_update: str):
    """1 baris: [mass update multi] [pricelist] [addon] [diskon manual]"""
    c1, c2, c3, c4 = st.columns([3.2, 2.2, 2.2, 1.4])

    with c1:
        templates = st.file_uploader(
            label_mass_update,
            type=["xlsx"],
            accept_multiple_files=True,
            key=f"{prefix}_tpl",
        )
    with c2:
        pricelist = st.file_uploader(
            "Upload Pricelist",
            type=["xlsx"],
            key=f"{prefix}_pl",
        )
    with c3:
        addon = st.file_uploader(
            "Upload Addon",
            type=["xlsx"],
            key=f"{prefix}_ad",
        )
    with c4:
        discount = st.number_input(
            "Diskon Manual",
            min_value=0,
            value=0,
            step=1000,
            key=f"{prefix}_disc",
        )

    return templates, pricelist, addon, int(discount)

def build_maps_price(spec: dict, pricelist_uploader, addon_uploader):
    pl_map = cached_build_pricelist_map(
        pricelist_bytes=pricelist_uploader.getvalue(),
        sheet_name=spec["pricelist"]["sheet_name"],
        header_row=spec["pricelist"]["header_row"],
        sku_header_candidates=tuple(spec["pricelist"]["sku_header_candidates"]),
        price_col_letter=spec["pricelist"]["price_col_letter"],
    )
    ad_map = cached_build_addon_map(
        addon_bytes=addon_uploader.getvalue(),
        code_candidates=tuple(spec["addon"]["code_candidates"]),
        price_candidates=tuple(spec["addon"]["price_candidates"]),
    )
    return pl_map, ad_map

def show_debug_block(debug_info: Dict, pricelist_len: Optional[int] = None, addon_len: Optional[int] = None):
    st.subheader("🔧 Debug Result")
    cols = st.columns(3)
    with cols[0]:
        st.write("Template:", debug_info.get("template"))
        if pricelist_len is not None:
            st.write("Pricelist SKU count:", int(pricelist_len))
        if addon_len is not None:
            st.write("Addon count:", int(addon_len))
    with cols[1]:
        st.write("rows_total_scanned:", debug_info.get("rows_total_scanned"))
        st.write("rows_with_sku:", debug_info.get("rows_with_sku", debug_info.get("rows_with_seller_sku")))
        st.write("rows_base_found:", debug_info.get("rows_base_found", debug_info.get("rows_stock_found")))
    with cols[2]:
        st.write("rows_changed:", debug_info.get("rows_changed"))
        st.write("rows_same_price/qty:", debug_info.get("rows_same_price", debug_info.get("rows_same_qty")))
        st.write("skipped_not_found:", debug_info.get("rows_skipped_base_not_found", debug_info.get("rows_skipped_stock_not_found")))

    if debug_info.get("sample_first_sku"):
        st.write("Contoh SKU (dari template):", debug_info["sample_first_sku"])

    if debug_info.get("sample_missing_base"):
        st.write("Contoh BASE SKU tidak ketemu di pricelist (maks 50):")
        st.dataframe(pd.DataFrame(debug_info["sample_missing_base"]), use_container_width=True)

    if debug_info.get("sample_missing_addon"):
        st.write("Contoh ADDON tidak ketemu di mapping (maks 50):")
        st.dataframe(pd.DataFrame(debug_info["sample_missing_addon"]), use_container_width=True)

    if debug_info.get("sample_missing_stock"):
        st.write("Contoh SKU tidak ketemu di file stok (maks 50):")
        st.dataframe(pd.DataFrame(debug_info["sample_missing_stock"]), use_container_width=True)


# =========================
# APP UI
# =========================
st.title("sellerengine")

debug_mode = st.checkbox("🔧 Debug Mode", value=False)

tab_hn, tab_hc, tab_st = st.tabs(["Harga Normal", "Harga Coret", "Update Stok"])

# =========================
# TAB: HARGA NORMAL
# =========================
with tab_hn:
    t1, t2, t3, t4 = st.tabs(["TikTok", "Shopee", "PowerMerchant", "BigSeller"])

    def harga_normal(platform: str, prefix: str, label: str):
        spec = SPECS[("harga_normal", platform)]
        templates, pricelist, addon, discount = ui_row_4(prefix, label)
        only_changed = st.checkbox("Hanya baris yang berubah", value=True, key=f"{prefix}_only")

        if st.button("PROCESS", type="primary", key=f"{prefix}_go"):
            if not templates or pricelist is None or addon is None:
                st.error("Wajib upload: File mass update + Pricelist + Addon.")
                st.stop()

            pl_map, ad_map = build_maps_price(spec, pricelist, addon)

            if debug_mode:
                st.info(
                    f"DEBUG: pricelist SKU={len(pl_map)} | addon={len(ad_map)} | "
                    f"note: harga pricelist auto x1000 (kalau < 1,000,000), diskon manual TIDAK x1000"
                )

            out_files = []
            all_debug = []
            prog = st.progress(0)

            for i, f in enumerate(templates):
                out_bytes, _changes, dbg = process_price_inplace(
                    template_bytes=f.getvalue(),
                    template_name=f.name,
                    spec=spec,
                    pricelist_map=pl_map,
                    addon_map=ad_map,
                    discount_rp=int(discount),
                    only_changed=only_changed,
                    debug=debug_mode,
                )
                if out_bytes:
                    out_files.append((f.name.replace(".xlsx", "_changed.xlsx"), out_bytes))
                if debug_mode:
                    all_debug.append(dbg)

                prog.progress((i + 1) / len(templates))

            download_outputs(out_files, f"harga_normal_{platform}.zip")

            if debug_mode and all_debug:
                st.divider()
                for dbg in all_debug[:3]:
                    show_debug_block(dbg, pricelist_len=len(pl_map), addon_len=len(ad_map))
                if len(all_debug) > 3:
                    st.warning(f"Debug ditampilkan hanya untuk 3 file pertama (total file: {len(all_debug)}).")

    with t1:
        harga_normal("tiktok", "hn_tt", "Upload File Mass Update (TikTok)")
    with t2:
        harga_normal("shopee", "hn_sp", "Upload File Mass Update (Shopee)")
    with t3:
        harga_normal("powermerchant", "hn_pm", "Upload File Mass Update (PowerMerchant)")
    with t4:
        harga_normal("bigseller", "hn_bs", "Upload File Mass Update (BigSeller)")

# =========================
# TAB: HARGA CORET
# =========================
with tab_hc:
    c1, c2, c3 = st.tabs(["TikTok (Discount Template)", "PowerMerchant (Discount Template)", "Shopee (Harga Diskon)"])

    # TikTok Discount Template
    with c1:
        platform = "tiktok"
        prefix = "hc_tt"
        spec = SPECS[("discount_template", platform)]

        templates, pricelist, addon, discount = ui_row_4(prefix, "Upload File Mass Update (TikTok Discount)")
        only_changed = st.checkbox("Hanya baris yang berubah", value=True, key=f"{prefix}_only")

        if st.button("PROCESS", type="primary", key=f"{prefix}_go"):
            if not templates or pricelist is None or addon is None:
                st.error("Wajib upload: File mass update + Pricelist + Addon.")
                st.stop()

            pl_map, ad_map = build_maps_price(spec, pricelist, addon)

            out_files, previews = [], []
            all_debug = []
            prog = st.progress(0)

            for i, f in enumerate(templates):
                files_out, df_prev, dbg = process_discount_template(
                    template_bytes=f.getvalue(),
                    template_name=f.name,
                    spec=spec,
                    pricelist_map=pl_map,
                    addon_map=ad_map,
                    discount_rp=int(discount),
                    only_changed=only_changed,
                    debug=debug_mode,
                )
                out_files.extend(files_out)
                if df_prev is not None and len(df_prev) > 0:
                    previews.append(df_prev)
                if debug_mode:
                    all_debug.append(dbg)
                prog.progress((i + 1) / len(templates))

            if previews:
                st.dataframe(pd.concat(previews, ignore_index=True).head(300), use_container_width=True)

            download_outputs(out_files, "tiktok_discount_output.zip")

            if debug_mode and all_debug:
                st.divider()
                for dbg in all_debug[:3]:
                    show_debug_block(dbg, pricelist_len=len(pl_map), addon_len=len(ad_map))
                if len(all_debug) > 3:
                    st.warning(f"Debug ditampilkan hanya untuk 3 file pertama (total file: {len(all_debug)}).")

    # PowerMerchant Discount Template (M4)
    with c2:
        platform = "powermerchant"
        prefix = "hc_pm"
        spec = SPECS[("discount_template", platform)]

        templates, pricelist, addon, discount = ui_row_4(prefix, "Upload File Mass Update (PowerMerchant Discount)")
        only_changed = st.checkbox("Hanya baris yang berubah", value=True, key=f"{prefix}_only")

        if st.button("PROCESS", type="primary", key=f"{prefix}_go"):
            if not templates or pricelist is None or addon is None:
                st.error("Wajib upload: File mass update + Pricelist + Addon.")
                st.stop()

            pl_map, ad_map = build_maps_price(spec, pricelist, addon)

            out_files, previews = [], []
            all_debug = []
            prog = st.progress(0)

            for i, f in enumerate(templates):
                files_out, df_prev, dbg = process_discount_template(
                    template_bytes=f.getvalue(),
                    template_name=f.name,
                    spec=spec,
                    pricelist_map=pl_map,
                    addon_map=ad_map,
                    discount_rp=int(discount),
                    only_changed=only_changed,
                    debug=debug_mode,
                )
                out_files.extend(files_out)
                if df_prev is not None and len(df_prev) > 0:
                    previews.append(df_prev)
                if debug_mode:
                    all_debug.append(dbg)
                prog.progress((i + 1) / len(templates))

            if previews:
                st.dataframe(pd.concat(previews, ignore_index=True).head(300), use_container_width=True)

            download_outputs(out_files, "powermerchant_discount_output.zip")

            if debug_mode and all_debug:
                st.divider()
                for dbg in all_debug[:3]:
                    show_debug_block(dbg, pricelist_len=len(pl_map), addon_len=len(ad_map))
                if len(all_debug) > 3:
                    st.warning(f"Debug ditampilkan hanya untuk 3 file pertama (total file: {len(all_debug)}).")

    # Shopee in-place Harga Diskon
    with c3:
        platform = "shopee"
        prefix = "hc_sp"
        spec = SPECS[("harga_coret", platform)]

        templates, pricelist, addon, discount = ui_row_4(prefix, "Upload File Mass Update (Shopee Harga Diskon)")
        only_changed = st.checkbox("Hanya baris yang berubah", value=True, key=f"{prefix}_only")

        if st.button("PROCESS", type="primary", key=f"{prefix}_go"):
            if not templates or pricelist is None or addon is None:
                st.error("Wajib upload: File mass update + Pricelist + Addon.")
                st.stop()

            pl_map, ad_map = build_maps_price(spec, pricelist, addon)

            out_files = []
            all_debug = []
            prog = st.progress(0)

            for i, f in enumerate(templates):
                out_bytes, _changes, dbg = process_price_inplace(
                    template_bytes=f.getvalue(),
                    template_name=f.name,
                    spec=spec,
                    pricelist_map=pl_map,
                    addon_map=ad_map,
                    discount_rp=int(discount),
                    only_changed=only_changed,
                    debug=debug_mode,
                )
                if out_bytes:
                    out_files.append((f.name.replace(".xlsx", "_changed.xlsx"), out_bytes))
                if debug_mode:
                    all_debug.append(dbg)
                prog.progress((i + 1) / len(templates))

            download_outputs(out_files, "harga_coret_shopee.zip")

            if debug_mode and all_debug:
                st.divider()
                for dbg in all_debug[:3]:
                    show_debug_block(dbg, pricelist_len=len(pl_map), addon_len=len(ad_map))
                if len(all_debug) > 3:
                    st.warning(f"Debug ditampilkan hanya untuk 3 file pertama (total file: {len(all_debug)}).")

# =========================
# TAB: UPDATE STOK
# =========================
with tab_st:
    s1, s2 = st.tabs(["TikTok", "Shopee"])

    def update_stok(platform: str, prefix: str, label: str):
        spec = SPECS[("update_stok", platform)]
        sheets_from = spec["stock_source"]["sheets_from"]
        sheets_to = spec["stock_source"]["sheets_to"]

        # Slot "Upload Pricelist" dipakai sebagai FILE STOK SUMBER (sheet LAPTOP..SER OTH CON)
        templates, stock_source_file, _addon_unused, _disc_unused = ui_row_4(prefix, label)

        mode = st.radio("Mode Stok", ["Nasional", "Area", "Toko"], horizontal=True, key=f"{prefix}_mode")

        qty_col = None
        df_stock = None

        if stock_source_file is not None:
            try:
                df_stock = cached_build_stock_dataframe_from_range(
                    stock_file_bytes=stock_source_file.getvalue(),
                    sheets_from=sheets_from,
                    sheets_to=sheets_to,
                )
                sku_cols = {"KODEBARANG", "KODE BARANG", "SKU"}
                qty_cols = [c for c in df_stock.columns if c not in sku_cols]
                qty_cols_sorted = sorted(qty_cols)

                default_nasional = None
                for c in ["TOT", "TOTAL", "NASIONAL"]:
                    if c in df_stock.columns:
                        default_nasional = c
                        break

                if mode == "Nasional":
                    qty_col = default_nasional
                    if qty_col is None:
                        st.error("Mode Nasional butuh kolom TOT/TOTAL/NASIONAL di file stok.")
                else:
                    if qty_cols_sorted:
                        qty_col = st.selectbox("Pilih Kolom Stok (Area/Toko)", qty_cols_sorted, key=f"{prefix}_qtycol")
                    else:
                        st.error("Tidak ada kolom stok yang bisa dipilih di file stok.")
            except Exception as e:
                st.error(f"Gagal baca file stok: {e}")

        if st.button("PROCESS", type="primary", key=f"{prefix}_go"):
            if not templates or stock_source_file is None:
                st.error("Wajib upload: File mass update + File stok sumber (pakai slot Upload Pricelist).")
                st.stop()

            if df_stock is None or qty_col is None:
                st.error("File stok belum siap / kolom stok belum valid.")
                st.stop()

            stock_map = build_stock_map_from_df(df_stock, qty_col=qty_col)

            out_files = []
            all_debug = []
            prog = st.progress(0)

            for i, f in enumerate(templates):
                out_bytes, _changes, dbg = process_stock_inplace(
                    template_bytes=f.getvalue(),
                    template_name=f.name,
                    spec=spec,
                    stock_value_map=stock_map,
                    debug=debug_mode,
                )
                if out_bytes:
                    out_files.append((f.name.replace(".xlsx", "_stok_changed.xlsx"), out_bytes))
                if debug_mode:
                    all_debug.append(dbg)
                prog.progress((i + 1) / len(templates))

            download_outputs(out_files, f"stok_{platform}.zip")

            if debug_mode and all_debug:
                st.divider()
                for dbg in all_debug[:3]:
                    show_debug_block(dbg)
                if len(all_debug) > 3:
                    st.warning(f"Debug ditampilkan hanya untuk 3 file pertama (total file: {len(all_debug)}).")

    with s1:
        update_stok("tiktok", "st_tt", "Upload File Mass Update (TikTok Stok)")
    with s2:
        update_stok("shopee", "st_sp", "Upload File Mass Update (Shopee Stok)")



